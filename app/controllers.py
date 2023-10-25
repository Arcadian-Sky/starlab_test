import os
import openpyxl
from datetime import datetime
from sqlalchemy import or_, and_
from sqlalchemy.ext.asyncio import AsyncSession
from sqlmodel import select
from app.db_connect import get_db_session
from app.models import *

file_upload_dir = 'upload'


async def get_data(book_filter: list):
    async with get_db_session() as db_session:
        query_authors = select(Author)
        result_authors = await db_session.execute(query_authors)
        authors = result_authors.scalars().all()

        serialized_authors = {author.id: {
            'id': author.id,
            'name': author.name,
            'second_name': author.second_name,
        } for author in authors}

        # Выборка по книгам
        query = select(Book)
        if book_filter:
            query = query.filter(and_(*book_filter))

        result = await db_session.execute(query)
        books = result.scalars().all()

        serialized_books = [{
            'id': book.id,
            'name': book.name,
            'genre': book.genre,
            'author': book.author_id,
            'file_path': book.file_path,
            'date_published': book.date_published.isoformat() if book.date_published else None
        } for book in books]

    return {
        'books': serialized_books,
        'authors': serialized_authors
    }


async def create_new_author(session: AsyncSession, name: str, second_name: str):
    try:
        new_row = Author(name=name, second_name=second_name)
        async with session.begin():
            session.add(new_row)
            await session.commit()
            return new_row.id
    except Exception as e:
        print(f"Ошибка при создании автора: {e}")
        return None


async def create_new_book(session: AsyncSession, name: str, author_id: int, date_published: str, genre: str,
                          file_data: bytes = None):
    try:
        if date_published:
            date_published = datetime.strptime(date_published, '%Y-%m-%d').date()

        new_book = Book(name=name, author_id=author_id, date_published=date_published, genre=genre)
        async with session.begin():
            session.add(new_book)
            await session.commit()

            if file_data is not None:
                async with get_db_session() as session:
                    query = select(Book).filter(Book.id == new_book.id)
                    result = await session.execute(query)
                    book = result.scalars().first()

                    book.file_path = file_upload_dir + '/' + str(new_book.id) + '/' + file_data.filename
                    await upload_file(file_upload_dir + '/' + str(new_book.id), file_data)
                    await session.commit()

        return new_book.id
    except Exception as e:
        print(f"Ошибка при создании книги: {e}")
        return None


async def upload_file(file_upload_path: str, file_field):
    current_directory = os.path.dirname(os.path.abspath(__file__))
    full_upload_path = os.path.join(current_directory, file_upload_path)

    if not os.path.exists(full_upload_path):
        os.makedirs(full_upload_path)

    if file_field and file_field.filename:
        file_path = os.path.join(full_upload_path, file_field.filename)
        with open(file_path, 'wb') as f:
            f.write(file_field.file.read())
        pass
    else:
        raise Exception('No file was sent in the request')


async def parse_and_decline_book(file_path: str):
    async with get_db_session() as db_session:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        author_ids = []
        books_names = []

        for sheet in workbook.worksheets:
            if sheet.title == 'author':
                author_ids.extend(await parse_authors(sheet, db_session))
            elif sheet.title == 'name':
                books_names.extend(await parse_books(sheet, db_session))

        book_filter = [Book.name.in_(books_names), Book.author_id.in_(author_ids)]
        query = select(Book).filter(or_(*book_filter))
        result = await db_session.execute(query)
        books = result.scalars().all()
        for book in books:
            book.file_path = None
            await db_session.commit()

        os.remove(file_path)
        return books

async def parse_authors(sheet, db_session):
    headers = get_headers(sheet)
    authors = []
    author_filters = []

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        name = row[headers.get('name')]  if 'name' in headers else None
        second_name = row[headers.get('second_name')]  if 'second_name' in headers else None
        if name or second_name:
            if name:
                author_filters.append(Author.name.like(f"%{name}%"))
            if second_name:
                author_filters.append(Author.second_name.like(f"%{second_name}%"))

    query = select(Author).filter(or_(*author_filters))
    result = await db_session.execute(query)
    authors.extend(result.scalars().all())

    author_ids = [author.id for author in authors]

    return author_ids

async def parse_books(sheet, db_session):
    headers = get_headers(sheet)
    book_names = []

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        name = row[headers.get('name')]
        if name:
            book_names.append(str(name))

    return book_names

def get_headers(sheet):
    headers = {}
    for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
        for idx, cell_value in enumerate(row):
            headers[cell_value] = idx
    return headers